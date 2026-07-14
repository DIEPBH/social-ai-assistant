from collections import Counter
from pathlib import Path
import re
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import unicodedata
from social_messages.models import IntakeSubmission, Report


class DailyReportExporter:
    def __init__(self, report: Report, requester_user=None, is_admin=False, is_specialist=False):
        self.report = report
        self.requester_user = requester_user
        self.is_admin = is_admin
        self.is_specialist = is_specialist

    def export(self) -> str:
        workbook = Workbook()

        summary_sheet = workbook.active
        summary_sheet.title = "Tổng quan"

        submission_sheet = workbook.create_sheet("Hồ sơ tiếp nhận")

        submissions = IntakeSubmission.objects.select_related(
            "conversation",
            "conversation__channel",
            "message",
        ).filter(
            created_at__gte=self.report.from_time,
            created_at__lte=self.report.to_time,
        ).exclude(
            message__sender_type="admin"
        )
        
        if self.is_specialist and not self.is_admin and self.requester_user:
            submissions = submissions.filter(assignments__user=self.requester_user).distinct()
            
        submissions = submissions.order_by("created_at")

        total_submissions = submissions.count()
        total_conversations = submissions.values("conversation_id").distinct().count()

        intent_counter = Counter()
        topic_counter = Counter()
        sentiment_counter = Counter()
        priority_counter = Counter()
        processing_status_counter = Counter()
        platform_counter = Counter()

        for submission in submissions:
            intent_counter[self._display_intent(submission.intent)] += 1
            processing_status_counter[self._display_processing_status(submission.processing_status)] += 1

            if submission.topic:
                topic_counter[submission.topic] += 1
            if submission.sentiment:
                sentiment_counter[submission.sentiment] += 1
            if submission.priority:
                priority_counter[self._display_priority(submission.priority)] += 1

            if submission.conversation and submission.conversation.channel:
                platform_counter[self._display_platform(submission.conversation.channel.platform)] += 1

        self._build_summary_sheet(
            summary_sheet=summary_sheet,
            total_submissions=total_submissions,
            total_conversations=total_conversations,
            intent_counter=intent_counter,
            processing_status_counter=processing_status_counter,
            platform_counter=platform_counter,
            topic_counter=topic_counter,
            sentiment_counter=sentiment_counter,
            priority_counter=priority_counter,
        )

        self._build_submission_sheet(submission_sheet, submissions)

        self._style_sheet(summary_sheet)
        self._style_sheet(submission_sheet)

        reports_dir = self._get_report_folder()
        reports_dir.mkdir(parents=True, exist_ok=True)

        safe_title = self._slugify_filename(self.report.title)
        filename = f"{safe_title}_{self.report.id}.xlsx"
        file_path = reports_dir / filename

        workbook.save(file_path)

        return str(file_path)

    def _build_summary_sheet(
        self,
        summary_sheet,
        total_submissions,
        total_conversations,
        intent_counter,
        processing_status_counter,
        platform_counter,
        topic_counter,
        sentiment_counter,
        priority_counter,
    ):
        summary_sheet.append(["Nội dung", "Giá trị"])
        summary_sheet.append(["Tiêu đề báo cáo", self.report.title])
        summary_sheet.append(["Từ thời gian", self.report.from_time.strftime("%d/%m/%Y %H:%M:%S")])
        summary_sheet.append(["Đến thời gian", self.report.to_time.strftime("%d/%m/%Y %H:%M:%S")])
        summary_sheet.append(["Tổng số hồ sơ tiếp nhận", total_submissions])
        summary_sheet.append(["Tổng số hội thoại có hồ sơ", total_conversations])
        summary_sheet.append([])

        self._append_counter(summary_sheet, "Thống kê theo loại hồ sơ", intent_counter)
        self._append_counter(summary_sheet, "Thống kê theo trạng thái xử lý", processing_status_counter)
        self._append_counter(summary_sheet, "Thống kê theo nền tảng", platform_counter)
        self._append_counter(summary_sheet, "Thống kê theo chủ đề", topic_counter)
        self._append_counter(summary_sheet, "Thống kê theo cảm xúc", sentiment_counter)
        self._append_counter(summary_sheet, "Thống kê theo mức ưu tiên", priority_counter)

    def _build_submission_sheet(self, sheet, submissions):
        sheet.append([
            "Mã hồ sơ",
            "Nền tảng",
            "Kênh tiếp nhận",
            "Mã hội thoại",
            "Loại hồ sơ",
            "Họ tên người gửi",
            "Số điện thoại",
            "Địa chỉ",
            "Nội dung tiếp nhận",
            "Thời gian xảy ra",
            "Địa điểm xảy ra",
            "Đối tượng liên quan",
            "Chủ đề",
            "Cảm xúc",
            "Mức ưu tiên",
            "Tóm tắt",
            "Nội dung phản hồi",
            "Trạng thái xử lý",
            "Thời gian tạo hồ sơ",
        ])

        for submission in submissions:
            channel = submission.conversation.channel

            sheet.append([
                submission.id,
                self._display_platform(channel.platform),
                channel.name,
                submission.conversation_id,
                self._display_intent(submission.intent),
                submission.citizen_name,
                submission.phone_number,
                submission.address,
                submission.content,
                submission.event_time,
                submission.event_location,
                submission.related_person,
                submission.topic,
                submission.sentiment,
                self._display_priority(submission.priority),
                submission.summary,
                submission.response_text,
                self._display_processing_status(submission.processing_status),
                submission.created_at.strftime("%d/%m/%Y %H:%M:%S"),
            ])

    def _append_counter(self, sheet, title, counter):
        sheet.append([])
        sheet.append([title, "Số lượng"])

        if not counter:
            sheet.append(["Không có dữ liệu", 0])
            return

        for label, count in counter.most_common():
            sheet.append([label, count])

    def _style_sheet(self, sheet):
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True)

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                value = cell.value
                if value is None:
                    continue

                value_length = len(str(value))
                if value_length > max_length:
                    max_length = value_length

            adjusted_width = min(max(max_length + 2, 14), 60)
            sheet.column_dimensions[column_letter].width = adjusted_width

        sheet.freeze_panes = "A2"

    def _display_platform(self, platform):
        mapping = {
            "zalo": "Zalo OA",
            "facebook": "Facebook Messenger",
        }
        return mapping.get(platform, platform or "")

    def _display_intent(self, intent):
        mapping = {
            "complaint": "Khiếu nại",
            "crime_report": "Tin báo tội phạm",
            "admin_procedure": "Hỏi thủ tục hành chính",
        }
        return mapping.get(intent, intent or "")

    def _slugify_filename(self, text):
        text = unicodedata.normalize("NFKD", text or "")
        text = text.encode("ascii", "ignore").decode("ascii")
        text = re.sub(r"[^\w\s-]", "", text)
        text = text.strip().replace(" ", "_")
        return text or "bao_cao"

    def _get_report_folder(self):
        target_time = self.report.from_time
        year = target_time.strftime("%Y")
        month = target_time.strftime("%m")
        return Path(settings.MEDIA_ROOT) / "reports" / year / month

    def _display_status(self, status):
        mapping = {
            "received": "Đã tiếp nhận",
            "validated": "Hợp lệ",
            "analyzed": "Đã phân tích",
            "responded": "Đã phản hồi",
            "rejected": "Không hợp lệ",
        }
        return mapping.get(status, status or "")

    def _display_priority(self, priority):
        mapping = {
            "low": "Thấp",
            "normal": "Bình thường",
            "medium": "Trung bình",
            "high": "Cao",
            "urgent": "Khẩn cấp",
        }
        return mapping.get(priority, priority or "")

    def _display_processing_status(self, status):
        mapping = {
            "unassigned": "Chưa phân công",
            "pending": "Chưa xử lý",
            "in_progress": "Đang xử lý",
            "completed": "Đã xử lý",
            "returned": "Trả lại",
        }
        return mapping.get(status, status or "")
    
    